from __future__ import annotations
from queue import Queue
import threading
import requests
import uuid
from collections import defaultdict
from django.core.cache import cache
import time 
import xml.etree.ElementTree as ET  
# from omnisight.decorators import redis_cache
from ecommerce_tool.util.marketplaces import get_filtered_marketplaces
from ecommerce_tool.util.shipping_price import get_full_order_and_shipping_details,get_orders_by_customer_and_date
from omnisight.operations.walmart_utils import getAccesstoken
from omnisight.models import *
import pandas as pd
from ecommerce_tool.crud import DatabaseModel
from mongoengine import DoesNotExist
import json
import pdfplumber
from rest_framework.parsers import JSONParser
from django.views.decorators.csrf import csrf_exempt
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO
from django.http import HttpResponse
from omnisight.operations.helium_utils import grossRevenue
from omnisight.operations.helium_utils import get_date_range, convertLocalTimeToUTC,convertdateTotimezone
import pytz
from pytz import timezone
import logging
from rest_framework.decorators import api_view
from rest_framework.response import Response    
import math
from django.http import JsonResponse
from rest_framework.parsers import JSONParser
from bson import ObjectId
from .helium_dashboard import sanitize_data,batch_get_sales_data_optimized,clean_json_floats
from threading import Thread
logger = logging.getLogger(__name__)
def sanitize_floats(data):
    """Recursively replace NaN, inf, -inf with None"""
    if isinstance(data, dict):
        return {k: sanitize_floats(v) for k, v in data.items()}
    elif isinstance(data, list):
        return [sanitize_floats(v) for v in data]
    elif isinstance(data, float):
        if math.isnan(data) or math.isinf(data):
            return None  
        return round(data, 2)  
    else:
        return data
    
def getMarketplaceList(request):
    country=request.GET.get('country')
    pipeline=[]
    if country:
        pipeline.append({
            "$match":{"country":country}
        })
    pipeline.append({
            "$project": {
                "_id": 0,
                "id": {"$toString": "$_id"},
                "name": 1,
                "image_url": 1,
                "fulfillment_channel": {
                    "$switch": {
                        "branches": [
                            {
                                "case": {"$eq": ["$name", "Amazon"]},
                                "then": [{"FBA": "AFN"}, {"FBM": "MFN"}]
                            },
                            # {
                            #     "case": {"$eq": ["$name", "Walmart"]},
                            #     "then": [{"FBM": "SellerFulfilled"}]
                            # }
                        ],
                        "default": []  
                    }
                }
            }
        })
    marketplace_list = list(Marketplace.objects.aggregate(*pipeline))
    return marketplace_list
@csrf_exempt
def getProductList(request):
    pipeline = [
            {
                "$project" : {
                    "_id" : 1,
                    "name" : 1,
                    "image_url" : 1,
                }
            }
        ]
    marketplace_list = list(Marketplace.objects.aggregate(*(pipeline)))
    data = dict()
    json_request = JSONParser().parse(request)
    marketplace_id = json_request.get('marketplace_id')
    country=json_request.get("country","US")
    filtered_marketplace_id=get_filtered_marketplaces(country,marketplace_id)
    skip = int(json_request.get('skip'))
    limit = int(json_request.get('limit'))
    search_query = json_request.get('search_query')   
    marketplace = json_request.get('marketplace')
    category_name = json_request.get('category_name')
    brand_id_list = json_request.get('brand_id_list')
    sort_by = json_request.get('sort_by')
    sort_by_value = json_request.get('sort_by_value')
    pipeline = []
    count_pipeline = []
    match = {}
    if filtered_marketplace_id and isinstance(filtered_marketplace_id,list) and len(filtered_marketplace_id)>0:
        match['marketplace_id'] = {"$in":filtered_marketplace_id}
    if category_name != None and category_name != "" and category_name != []:
        match['category'] = {"$in":category_name}
    if brand_id_list != None and brand_id_list != "" and brand_id_list != []:
        match['brand_id'] = {"$in":[ObjectId(brand_id) for brand_id in brand_id_list]}
    if search_query != None and search_query != "":
        search_query = re.escape(search_query.strip())
        match["$or"] = [
            {"product_title": {"$regex": search_query, "$options": "i"}},
            {"sku": {"$regex": search_query, "$options": "i"}},
        ]
    if match != {}:
        match_pipeline = {
            "$match" : match}
        pipeline.append(match_pipeline)
        count_pipeline.append(match_pipeline)
    pipeline.extend([
        {
            "$project" : {
                "_id" : 0,
                "id" : {"$toString" : "$_id"},
                "product_title" : 1,
                "product_id" : 1,
                "sku" : 1,
                "asin" : {"$ifNull" : ["$asin",""]},  
                "price" : 1,
                "quantity" : 1,
                "published_status" : 1,
                "category" : {"$ifNull" : ["$category",""]},  
                "image_url" : {"$ifNull" : ["$image_url",""]},  
                "marketplace_ids": {"$ifNull": ["$marketplace_ids", []]},  
                "marketplace_ins": [],
                "marketplace_image_url": [] 
            }
        },
        {
            "$skip" : skip
        },
        {
            "$limit" : limit  
        }
    ])
    if sort_by != None and sort_by != "":
        sort = {
            "$sort" : {
                sort_by : int(sort_by_value)
            }
        }
        pipeline.append(sort)
    product_list = list(Product.objects.aggregate(*(pipeline)))
    for ins in product_list:
        marketplace_ids = ins.get('marketplace_ids', [])
        ins['marketplace_details'] = []
        for marketplace in marketplace_list:
            if marketplace['_id'] in marketplace_ids:
                ins['marketplace_ins'].append(marketplace['name'])
                ins['marketplace_image_url'].append(marketplace['image_url'])
        del ins['marketplace_ids']
    count_pipeline.extend([
        {
            "$count": "total_count"
        }
    ])
    total_count_result = list(Product.objects.aggregate(*(count_pipeline)))
    total_count = total_count_result[0]['total_count'] if total_count_result else 0
    data['total_count'] = total_count
    data['product_list'] = product_list
    return data

def getProductCategoryList(request):
    data = dict()
    marketplace_id = request.GET.get('marketplace_id')
    match = {}
    if marketplace_id != None and marketplace_id != "":
        match['marketplace_id'] = ObjectId(marketplace_id)
    match['end_level'] = True
    pipeline = [
        {
            "$match": match    
        },
        {
            "$lookup": {
                "from": "product",
                "localField": "name",
                "foreignField": "category",
                "as": "products"
            }
        },
        {
            "$project": {
                "_id": 0,
                "id": {"$toString": "$_id"},
                "name": 1,
                "product_count": {"$size": "$products"}
            }
        },
        {
            "$match": {
                "product_count": {"$gt": 0}
            }
        }
    ]
    category_list = list(Category.objects.aggregate(*(pipeline)))
    data['category_list'] = category_list
    return data
def getBrandList(request):
    data = dict()
    marketplace_id = request.GET.get('marketplace_id')
    pipeline = []
    if marketplace_id != None and marketplace_id != "":
        match = {
            "$match" : {
                "marketplace_id" : ObjectId(marketplace_id)
            }
        }
        pipeline.append(match)
    pipeline.extend([
        {
            "$lookup": {
                "from": "product",
                "localField": "_id",
                "foreignField": "brand_id",
                "as": "products"
            }
        },
        {
            "$project" : {
                "_id" : 0,
                "id" : {"$toString" : "$_id"},
                "name" : 1,
                "product_count": {"$size": "$products"}
            }
        }
    ])
    brand_list = list(Brand.objects.aggregate(*(pipeline)))
    data['brand_list'] = brand_list
    return data
def fetchProductDetails(request):
    data = dict()
    product_id = request.GET.get('product_id')
    pipeline = [
        {
            "$match": {
                "_id": ObjectId(product_id)
            }
        },
        {
            "$lookup": {
                "from": "marketplace",
                "localField": "marketplace_ids",
                "foreignField": "_id",
                "as": "marketplace_ins"
            }
        },
        {
            "$project": {
                "_id": 0,
                "id": {"$toString": "$_id"},
                "product_title": {"$ifNull": ["$product_title", ""]},
                "product_description": {"$ifNull": ["$product_description", ""]},
                "product_id": {"$ifNull": ["$product_id", ""]},
                "product_id_type": {"$ifNull": ["$product_id_type", ""]},
                "sku": {"$ifNull": ["$sku", ""]},
                "price": {"$round": [{"$ifNull": ["$price", 0]}, 2]},
                "currency": {"$ifNull": ["$currency", ""]},
                "quantity": {"$ifNull": ["$quantity", 0]},
                "published_status": 1,
                "marketplace_ins": {
                    "$map": {
                        "input": "$marketplace_ins",
                        "as": "marketplace",
                        "in": "$$marketplace.name"
                    }
                },
                "marketplace_image_url": {
                    "$map": {
                        "input": "$marketplace_ins",
                        "as": "marketplace",
                        "in": "$$marketplace.image_url"
                    }
                },
                "item_condition": {"$ifNull": ["$item_condition", ""]},
                "item_note": {"$ifNull": ["$item_note", ""]},
                "listing_id": {"$ifNull": ["$listing_id", ""]},
                "upc": {"$ifNull": ["$upc", ""]},
                "gtin": {"$ifNull": ["$gtin", ""]},
                "asin": {"$ifNull": ["$asin", ""]},
                "model_number": {"$ifNull": ["$model_number", ""]},
                "category": {"$ifNull": ["$category", ""]},
                "brand_name": {"$ifNull": ["$brand_name", ""]},
                "manufacturer_name": {"$ifNull": ["$manufacturer_name", ""]},
                "attributes": {"$ifNull": ["$attributes", {}]},
                "features": {"$ifNull": ["$features", []]},
                "shelf_path": {"$ifNull": ["$shelf_path", ""]},
                "image_url": {"$ifNull": ["$image_url", ""]},
                "image_urls": {"$ifNull": ["$image_urls", []]},
                "vendor_funding": {"$round": [{"$ifNull": ["$vendor_funding", 0.0]}, 2]},
                "vendor_discount": {"$round": [{"$ifNull": ["$vendor_discount", 0.0]}, 2]},
                "product_cost": {"$round": [{"$ifNull": ["$product_cost", 0]}, 2]},
                "referral_fee": {"$round": [{"$ifNull": ["$referral_fee", 0]}, 2]},
                "a_shipping_cost": {"$round": [{"$ifNull": ["$a_shipping_cost", 0]}, 2]},
                "total_cogs": {"$round": [{"$ifNull": ["$total_cogs", 0]}, 2]},
                "w_product_cost": {"$round": [{"$ifNull": ["$w_product_cost", 0]}, 2]},
                "walmart_fee": {"$round": [{"$ifNull": ["$walmart_fee", 0]}, 2]},
                "w_shiping_cost": {"$round": [{"$ifNull": ["$w_shiping_cost", 0]}, 2]},
                "w_total_cogs": {"$round": [{"$ifNull": ["$w_total_cogs", 0]}, 2]},
                "pack_size": {"$ifNull": ["$pack_size", ""]},
            }
        }
    ]
    product_details = list(Product.objects.aggregate(*(pipeline)))
    if len(product_details):
        data = product_details[0]
    return data
def getOrdersBasedOnProduct(request):
    data = {'total_count': 0, 'orders': []}
    product_id = request.GET.get('product_id')
    skip = int(request.GET.get('skip', 0))
    limit = int(request.GET.get('limit', 10))
    sort_by = request.GET.get('sort_by')
    sort_by_value = int(request.GET.get('sort_by_value', -1)) 
    matching_order_items = list(OrderItems.objects.filter(ProductDetails__product_id=ObjectId(product_id)).only('id'))
    if not matching_order_items:
        return data
    total_count = Order.objects.filter(order_items__in=matching_order_items).count()
    sort_field = sort_by if sort_by else "order_date"
    sort_order = sort_by_value if sort_by_value else -1
    orders_queryset = Order.objects.filter(order_items__in=matching_order_items)\
            .order_by(f"{sort_order}{sort_field}")\
            .skip(skip).limit(limit - skip)
    order_list = list(orders_queryset)
    marketplace_ids = [o.marketplace_id.id for o in order_list if o.marketplace_id]
    marketplaces = {
        m.id: m.name for m in Marketplace.objects.filter(id__in=marketplace_ids)
    }
    orders = []
    for o in order_list:
        orders.append({
            "id": str(o.id),
            "purchase_order_id": o.purchase_order_id,
            "order_date": o.order_date.strftime('%Y-%m-%dT%H:%M:%S.%fZ')[:-3],
            "order_status": o.order_status,
            "order_total": round(o.order_total, 2) if o.order_total else 0.0,
            "currency": o.currency,
            "marketplace_name": marketplaces.get(o.marketplace_id.id, None),
        })
    custom_orders = list(custom_order.objects.filter(
        ordered_products__product_id =  ObjectId(product_id)
    ).only("order_id", "purchase_order_date", "order_status", "total_price", "currency"))
    for c in custom_orders:
        orders.append({
            "id": str(c.id),
            "purchase_order_id": c.order_id,
            "order_date": c.purchase_order_date.strftime('%Y-%m-%dT%H:%M:%S.%fZ')[:-3],
            "order_status": c.order_status,
            "order_total": round(c.total_price, 2) if c.total_price else 0.0,
            "currency": c.currency or "USD",
            "marketplace_name": "custom"
        })
    data = {
        "total_count": total_count,
        "orders": orders
    }
    return data
@csrf_exempt
@api_view(["POST"])
def fetchAllorders(request):
    data = dict()
    pipeline = []
    count_pipeline = []
    json_request = request.data
    user_id = json_request.get('user_id')
    limit = int(json_request.get('limit', 100)) 
    skip = int(json_request.get('skip', 0))  
    marketplace_id = json_request.get('marketplace_id')
    country = json_request.get("country", "US")
    filtered_marketplace_id = get_filtered_marketplaces(country, marketplace_id)
    sort_by = json_request.get('sort_by')
    sort_by_value = json_request.get('sort_by_value')
    search_query = json_request.get('search_query')
    order_status_filter = json_request.get('order_status') 

    if search_query and search_query.strip():
        skip = 0
        if not limit or limit > 100:
            limit = 25

    pacific_tz = pytz.timezone("US/Pacific")
    current_time_pacific = datetime.now(pacific_tz)

    if marketplace_id == "custom":
        base_pipeline = []
        if search_query and search_query.strip():
            safe_search = re.escape(search_query.strip())
            base_pipeline.append({"$match": {"order_id": {"$regex": safe_search, "$options": "i"}}})
        count_pipeline = base_pipeline + [{"$count": "total_count"}]
        total_count_result = list(custom_order.objects.aggregate(*count_pipeline))
        total_count = total_count_result[0]['total_count'] if total_count_result else 0
        pipeline = base_pipeline + [
            {"$project": {
                "_id": 0, "id": {"$toString": "$_id"}, "order_id": {"$ifNull": ["$order_id", ""]},
                "customer_name": {"$ifNull": ["$customer_name", ""]}, "shipping_address": {"$ifNull": ["$shipping_address", ""]},
                "total_quantity": {"$ifNull": ["$total_quantity", 0]}, "total_price": {"$ifNull": [{"$round": ["$total_price", 2]}, 0.0]},
                "purchase_order_date": {"$ifNull": ["$purchase_order_date", None]}, "expected_delivery_date": {"$ifNull": ["$expected_delivery_date", None]},
                "order_status" : "$order_status", "currency" : {"$ifNull" : ["$currency","USD"]}
            }}
        ]
        if sort_by:
            pipeline.append({"$sort": {sort_by: int(sort_by_value)}})
        else:
            pipeline.append({"$sort": {"id": -1}})
        pipeline.extend([{"$skip": skip}, {"$limit": limit}])
        data['manual_orders'] = list(custom_order.objects.aggregate(*pipeline))
        data['total_count'] = total_count
        data['status'] = "custom"
    else:
        # Add marketplace filter
        if filtered_marketplace_id and isinstance(filtered_marketplace_id, list) and len(filtered_marketplace_id) > 0:
            pipeline.append({"$match": {"marketplace_id": {"$in": filtered_marketplace_id}}})
        
        if search_query and search_query.strip():
            safe_search = re.escape(search_query.strip())
            pipeline.append({"$match": {"purchase_order_id": {"$regex": safe_search, "$options": "i"}}})
        
        if order_status_filter and order_status_filter != 'all':
            pipeline.append({"$match": {"order_status": order_status_filter}})
        
        count_pipeline = pipeline + [{"$count": "total_count"}]
        total_count_result = list(Order.objects.aggregate(*count_pipeline))
        total_count = total_count_result[0]['total_count'] if total_count_result else 0
        
        if sort_by:
            pipeline.append({"$sort": {sort_by: int(sort_by_value)}})
        else:
            pipeline.append({"$sort": {"order_date": -1}})
        
        pipeline.extend([
            {"$skip": skip}, {"$limit": limit},
            {"$lookup": {"from": "marketplace", "localField": "marketplace_id", "foreignField": "_id", "as": "marketplace_ins"}},
            {"$unwind": "$marketplace_ins"},
            {"$project": {
                "_id": 0, "id": {"$toString": "$_id"}, "purchase_order_id": 1,
                "order_date": 1, "order_status": 1, "order_total": 1, "currency": 1,
                "marketplace_name": "$marketplace_ins.name", "items_order_quantity": 1
            }}
        ])
        
        orders = list(Order.objects.aggregate(*(pipeline)))
        for order in orders:
            if order.get('order_date'):
                order_date_utc = order['order_date'].astimezone(pytz.utc) 
                order['order_date'] = order_date_utc.astimezone(pacific_tz)  
        orders = [o for o in orders if o.get('order_date') and o['order_date'] <= current_time_pacific]
        data['orders'] = orders
        data['total_count'] = total_count
        data['status'] = ""

    marketplace_pipeline = [{"$project" : {"_id" : 0, "id" : {"$toString" : "$_id"}, "name" : 1, "image_url" : 1}}]
    data['marketplace_list'] = list(Marketplace.objects.aggregate(*marketplace_pipeline))
    return Response(data)
def fetchOrderDetails(request):
    data = {}
    user_id = request.GET.get('user_id')
    order_id = request.GET.get('order_id')
    pipeline = [
        {
            "$match": {
                "_id": ObjectId(order_id)
            }
        },
        {
            "$lookup": {
                "from": "marketplace",
                "localField": "marketplace_id",
                "foreignField": "_id",
                "as": "marketplace_ins"
            }
        },
        {
            "$unwind": "$marketplace_ins"
        },
        {
            "$project": {
                "_id": 0,
                "id": {"$toString": "$_id"},
                "purchase_order_id": {"$ifNull": ["$purchase_order_id", ""]},
                "customer_order_id": {"$ifNull": ["$customer_order_id", ""]},
                "seller_order_id": {"$ifNull": ["$seller_order_id", ""]},
                "customer_email_id": {"$ifNull": ["$customer_email_id", ""]},
                "shipping_price":{"$ifNull":["$shipping_price",0]},
                "order_date": {
                    "$dateToString": {
                        "format": "%Y-%m-%dT%H:%M:%S.%LZ",
                        "date": {"$ifNull": ["$order_date", None]},
                    }
                },
                "earliest_ship_date": {
                    "$dateToString": {
                        "format": "%Y-%m-%dT%H:%M:%S.%LZ",
                        "date": {"$ifNull": ["$earliest_ship_date", None]},
                    }
                },
                "latest_ship_date": {
                    "$dateToString": {
                        "format": "%Y-%m-%dT%H:%M:%S.%LZ",
                        "date": {"$ifNull": ["$latest_ship_date", None]},
                    }
                },
                "last_update_date": {
                    "$dateToString": {
                        "format": "%Y-%m-%dT%H:%M:%S.%LZ",
                        "date": {"$ifNull": ["$last_update_date", None]},
                    }
                },
                "shipping_information": {"$ifNull": ["$shipping_information", {}]},
                "ship_service_level": {"$ifNull": ["$ship_service_level", ""]},
                "shipment_service_level_category": {"$ifNull": ["$shipment_service_level_category", ""]},
                "automated_shipping_settings": {"$ifNull": ["$automated_shipping_settings", {}]},
                "order_status": {"$ifNull": ["$order_status", ""]},
                "number_of_items_shipped": {"$ifNull": ["$number_of_items_shipped", 0]},
                "number_of_items_unshipped": {"$ifNull": ["$number_of_items_unshipped", 0]},
                "fulfillment_channel": {"$ifNull": ["$fulfillment_channel", ""]},
                "sales_channel": {"$ifNull": ["$sales_channel", ""]},
                "order_type": {"$ifNull": ["$order_type", ""]},
                "is_premium_order": {"$ifNull": ["$is_premium_order", False]},
                "is_prime": {"$ifNull": ["$is_prime", False]},
                "has_regulated_items": {"$ifNull": ["$has_regulated_items", False]},
                "is_replacement_order": {"$ifNull": ["$is_replacement_order", False]},
                "is_sold_by_ab": {"$ifNull": ["$is_sold_by_ab", False]},
                "is_ispu": {"$ifNull": ["$is_ispu", False]},
                "is_access_point_order": {"$ifNull": ["$is_access_point_order", False]},
                "is_business_order": {"$ifNull": ["$is_business_order", False]},
                "marketplace_name": {"$ifNull": ["$marketplace_ins.name", ""]},
                "payment_method": {"$ifNull": ["$payment_method", ""]},
                "payment_method_details": {"$ifNull": ["$payment_method_details", []]},
                "order_total": {"$ifNull": [{"$round": ["$order_total", 2]}, 0]},
                "currency": {"$ifNull": ["$currency", ""]},
                "is_global_express_enabled": {"$ifNull": ["$is_global_express_enabled", False]},
                "order_items": {"$ifNull": ["$order_items", []]},
                "merchant_order_id": {"$ifNull": ["$merchant_order_id", ""]}, 
                "merchant_shipment_cost":{"$ifNull":["$merchant_shipment_cost",0]}
            }
        }
    ]
    order_number = None
    merchant_shipment_cost=0
    order_details = list(Order.objects.aggregate(*pipeline))
    if order_details:
        data = order_details[0]
        fulfillment_channel = data.get("fulfillment_channel", "")
        merchant_shipment_cost=data.get('merchant_shipment_cost',0)
        if not merchant_shipment_cost:
            if fulfillment_channel=='SellerFulfilled':
                customer_email=data.get('customer_email_id','')
                order_date=data.get('order_date',None)
                po_id=data.get('purchase_order_id',"")
                shipping_info=get_orders_by_customer_and_date(
                    customer_email=customer_email,
                    order_date_utc_iso=order_date,
                    purchase_order_id=po_id,
                    local_tz='US/Pacific'
                )
                shipment_cost=0
                if shipping_info:
                    shipment_cost=float(shipping_info[-1].get('shipmentCost',0) or 0)
                order_obj=Order.objects(id=ObjectId(order_id)).first()
                if order_obj:
                    order_obj.update(set__merchant_shipment_cost=shipment_cost)
                data['merchant_shipment_cost']=shipment_cost
            elif fulfillment_channel == "MFN":
                order_number = data.get("merchant_order_id") or data.get("customer_order_id")
            if order_number:
                full_order_details = get_full_order_and_shipping_details(order_number)
                shipment_cost = 0
                if full_order_details and full_order_details.get('shipments'):
                    shipment_cost = float(full_order_details['shipments'][-1].get('shipmentCost', 0) or 0)
                    order_obj = Order.objects(id=ObjectId(order_id)).first()
                    if order_obj:
                        order_obj.update(set__merchant_shipment_cost=shipment_cost)
                        data["merchant_shipment_cost"] = shipment_cost
        order_items_ids = data.get("order_items", [])
        if order_items_ids:
            order_items = DatabaseModel.list_documents(OrderItems.objects, filters={"id__in": order_items_ids})
            serialized_items = []
            for item in order_items:
                serialized_item = {
                    "id": str(item.id),
                    "OrderId": item.OrderId,
                    "Platform": item.Platform,
                    "ProductDetails": {
                        "product_id": str(item.ProductDetails.product_id.id) if item.ProductDetails and item.ProductDetails.product_id else None,
                        "Title": item.ProductDetails.Title if item.ProductDetails else None,
                        "SKU": item.ProductDetails.SKU if item.ProductDetails else None,
                        "Condition": item.ProductDetails.Condition if item.ProductDetails else None,
                        "QuantityOrdered": item.ProductDetails.QuantityOrdered if item.ProductDetails else 0,
                        "QuantityShipped": item.ProductDetails.QuantityShipped if item.ProductDetails else 0,
                        "unit_price": float(item.ProductDetails.product_id.price) if item.ProductDetails and item.ProductDetails.product_id and item.ProductDetails.product_id.price else 0.0,
                    },
                    "Pricing": item.Pricing.to_mongo() if hasattr(item.Pricing, "to_mongo") else (dict(item.Pricing) if item.Pricing else {}),
                    "Fulfillment": {
                        "FulfillmentOption": item.Fulfillment.FulfillmentOption if item.Fulfillment else None,
                        "ShipMethod": item.Fulfillment.ShipMethod if item.Fulfillment else None,
                        "Carrier": item.Fulfillment.Carrier if item.Fulfillment else None,
                        "TrackingNumber": item.Fulfillment.TrackingNumber if item.Fulfillment else None,
                        "TrackingURL": item.Fulfillment.TrackingURL if item.Fulfillment else None,
                        "ShipDateTime": item.Fulfillment.ShipDateTime.isoformat() if item.Fulfillment and item.Fulfillment.ShipDateTime else None
                    },
                    "OrderStatus": {
                        "Status": item.OrderStatus.Status if item.OrderStatus else None,
                        "StatusDate": item.OrderStatus.StatusDate.isoformat() if item.OrderStatus and item.OrderStatus.StatusDate else None
                    },
                    "TaxCollection": item.TaxCollection.to_mongo() if hasattr(item.TaxCollection, "to_mongo") else (dict(item.TaxCollection) if item.TaxCollection else {}),
                    "IsGift": item.IsGift
                }
                serialized_items.append(serialized_item)
            data["order_items"] = serialized_items
        else:
            data["order_items"] = []
    return data
@csrf_exempt
def getProductListForOrdercreation(request):
    data = dict()
    json_request = JSONParser().parse(request)
    marketplace_id = json_request.get('marketplace_id')
    skip = int(json_request.get('skip',0))
    limit = int(json_request.get('limit',50))
    search_query = json_request.get('search_query')
    pipeline = []
    count_pipeline = []
    match = {}
    if marketplace_id != None and marketplace_id != "":
        match['marketplace_id'] = ObjectId(marketplace_id)
    if search_query != None and search_query != "":
        match['product_title'] = {"$regex": search_query, "$options": "i"}
    if match != {}:
        match_pipeline = {
            "$match" : match}
        pipeline.append(match_pipeline)
    pipeline.extend([
        {
            "$project" : {
                "_id" : 0,
                "id" : {"$toString" : "$_id"},
                "product_title" : 1,
                "sku" : 1,
                "price" : 1,
            }
        },
        {
            "$skip" : skip
        },
        {
            "$limit" : limit
        }
    ])
    product_list = list(Product.objects.aggregate(*(pipeline)))
    count_pipeline.extend([
        {
            "$count": "total_count"
        }
    ])
    total_count_result = list(Product.objects.aggregate(*(count_pipeline)))
    total_count = total_count_result[0]['total_count'] if total_count_result else 0
    data['total_count'] = total_count
    return product_list
@csrf_exempt
def createManualOrder(request):
    data = dict()
    json_request = JSONParser().parse(request)
    product_detail = list()
    ordered_products = json_request.get('ordered_products')
    custom_product_obj = json_request.get('custom_product_obj')
    user_id = json_request.get('user_id')
    if user_id:
        custom_product_obj['user_id'] = ObjectId(user_id)
    custom_product_obj['order_id'] = str(''.join([str(uuid.uuid4().int)[:13]]))
    custom_product_obj['customer_order_id'] = datetime.now().strftime('%Y%m%d') + str(uuid.uuid4().int)[:5]
    try:
        custom_product_obj['purchase_order_date'] = datetime.strptime(
            json_request.get(custom_product_obj['purchase_order_date']), '%Y-%m-%d'
        )
    except:
        pass
    expected_delivery_date = custom_product_obj.get('expected_delivery_date')
    if expected_delivery_date:
        custom_product_obj['expected_delivery_date'] = datetime.strptime(expected_delivery_date, '%Y-%m-%d')
    else:
        custom_product_obj['expected_delivery_date'] = None
    for ins in ordered_products:
        product_detail_dict = product_details(
            product_id=ObjectId(ins.get('product_id')),
            title=ins.get('title'),
            sku=ins.get('sku'),
            unit_price=float(ins.get('unit_price', 0)),
            quantity=int(ins.get('quantity', 0)),
            quantity_price=float(ins.get('quantity_price', 0))
        )
        product_detail.append(product_detail_dict)
    custom_product_obj['ordered_products'] = product_detail
    total_price = float(custom_product_obj.get('total_price', 0))
    discount = float(custom_product_obj.get('discount', 0))  
    tax = float(custom_product_obj.get('tax', 0))  
    shipment_cost = float(custom_product_obj.get('shipment_cost', 0))
    if discount > 0:
        custom_product_obj['discount_amount'] = (total_price * discount / 100)  
        total_price -= custom_product_obj['discount_amount']
    if tax > 0:
        custom_product_obj['tax_amount'] = (total_price * tax / 100)  
        total_price += custom_product_obj['tax_amount']
    custom_product_obj['total_price'] = round(total_price + shipment_cost, 2)  
    manual_order = DatabaseModel.save_documents(custom_order, custom_product_obj)
    data['message'] = "Manual order created successfully."
    data['order_id'] = str(manual_order.id)
    return data
@csrf_exempt
def listManualOrders(request):
    data = dict()
    json_request = JSONParser().parse(request)
    limit = int(json_request.get('limit', 100))  
    skip = int(json_request.get('skip', 0))  
    sort_by = json_request.get('sort_by')
    sort_by_value = json_request.get('sort_by_value')
    pipeline = [
        {
            "$project": {
                "_id": 0,
                "id": {"$toString": "$_id"},
                "order_id": {"$ifNull": ["$order_id", ""]},
                "customer_name": {"$ifNull": ["$customer_name", ""]},
                "shipping_address": {"$ifNull": ["$shipping_address", ""]},
                "total_quantity": {"$ifNull": ["$total_quantity", 0]},
                "total_price": {"$ifNull": [{"$round": ["$total_price", 0.0]}, 0.0]},
                "purchase_order_date": {"$ifNull": ["$purchase_order_date", None]},
                "expected_delivery_date": {"$ifNull": ["$expected_delivery_date", None]},
            }
        },
        {
            "$skip": skip
        },
        {
            "$limit": limit + skip
        }
    ]
    if sort_by != None and sort_by != "":
        sort = {
            "$sort" : {
                sort_by : int(sort_by_value)
            }
        }
        pipeline.append(sort)
    manual_orders = list(custom_order.objects.aggregate(*pipeline))
    count_pipeline = [
        {
            "$count": "total_count"
        }
    ]
    total_count_result = list(custom_order.objects.aggregate(*(count_pipeline)))
    total_count = total_count_result[0]['total_count'] if total_count_result else 0
    data['total_count'] = total_count
    data['manual_orders'] = manual_orders
    return data
@csrf_exempt
def updateManualOrder(request):
    data = dict()
    json_request = JSONParser().parse(request)
    order_id = json_request.get('order_id')
    ordered_products = json_request.get('ordered_products',[])
    custom_product_obj = json_request.get('custom_product_obj')
    try:
        purchase_order_date = custom_product_obj.get('purchase_order_date')
        if purchase_order_date:
            custom_product_obj['purchase_order_date'] = datetime.strptime(purchase_order_date, '%Y-%m-%d')
    except:
        pass
    expected_delivery_date = custom_product_obj.get('expected_delivery_date')
    if expected_delivery_date:
        custom_product_obj['expected_delivery_date'] = datetime.strptime(expected_delivery_date, '%Y-%m-%d')
    else:
        custom_product_obj['expected_delivery_date'] = None
    product_detail = []
    for ins in ordered_products:
        product_detail_dict = product_details(
            product_id=ObjectId(ins.get('product_id')),
            title=ins.get('title'),
            sku=ins.get('sku'),
            unit_price=float(ins.get('unit_price', 0)),
            quantity=int(ins.get('quantity', 0)),
            quantity_price=float(ins.get('quantity_price', 0))
        )
        product_detail.append(product_detail_dict)
    custom_product_obj['ordered_products'] = product_detail
    total_price = float(custom_product_obj.get('total_price', 0))
    discount = float(custom_product_obj.get('discount', 0))  
    tax = float(custom_product_obj.get('tax', 0))  
    shipment_cost = float(custom_product_obj.get('shipment_cost', 0))
    if discount > 0:
        custom_product_obj['discount_amount'] = (total_price * discount / 100)  
        total_price -= custom_product_obj['discount_amount']
    if tax > 0:
        custom_product_obj['tax_amount'] = (total_price * tax / 100)  
        total_price += custom_product_obj['tax_amount']
    custom_product_obj['total_price'] = round(total_price + shipment_cost, 2)  
    DatabaseModel.update_documents(custom_order.objects,{"id" : order_id},custom_product_obj)
    data['message'] = "Manual order updated successfully."
    return data
def fetchManualOrderDetails(request):
    data = dict()
    order_id = request.GET.get('order_id')
    pipeline = [
        {
            "$match": {
                "_id": ObjectId(order_id)
            }
        },
        {
            "$lookup": {
                "from": "product",
                "localField": "ordered_products.product_id",
                "foreignField": "_id",
                "as": "product_details"
            }
        },
        {
            "$project": {
                "_id": 0,
                "id": {"$toString": "$_id"},
                "order_id": {"$ifNull": ["$order_id", ""]},
                "customer_name": {"$ifNull": ["$customer_name", ""]},
                "shipping_address": {"$ifNull": ["$shipping_address", ""]},
                "total_quantity": {"$ifNull": ["$total_quantity", 0]},
                "total_price": {"$ifNull": ["$total_price", 0.0]},
                "shipment_type": {"$ifNull": ["$shipment_type", ""]},
                "channel": {"$ifNull": ["$channel", ""]},
                "order_status": {"$ifNull": ["$order_status", "Pending"]},
                "payment_status": {"$ifNull": ["$payment_status", "Pending"]},
                "payment_mode": {"$ifNull": ["$payment_mode", ""]},
                "invoice": {"$ifNull": ["$invoice", ""]},
                "transaction_id": {"$ifNull": ["$transaction_id", ""]},
                "tax": {"$ifNull": ["$tax", 0.0]},
                "tax_amount": {"$ifNull": [{"$round":["$tax_amount",2]}, 0.0]},
                "discount_amount": {"$ifNull": [{"$round":["$discount_amount",2]}, 0.0]},
                "discount": {"$ifNull": ["$discount", 0.0]},
                "supplier_name": {"$ifNull": ["$supplier_name", ""]},
                "mail": {"$ifNull": ["$mail", ""]},
                "contact_number": {"$ifNull": ["$contact_number", ""]},
                "customer_note": {"$ifNull": ["$customer_note", ""]},
                "tags": {"$ifNull": ["$tags", ""]},
                "package_dimensions": {"$ifNull": ["$package_dimensions", ""]},
                "weight": {"$ifNull": ["$weight", 0.0]},
                "shipment_cost": {"$ifNull": ["$shipment_cost", 0.0]},
                "shipment_speed": {"$ifNull": ["$shipment_speed", ""]},
                "shipment_mode": {"$ifNull": ["$shipment_mode", ""]},
                "carrier": {"$ifNull": ["$carrier", ""]},
                "tracking_number": {"$ifNull": ["$tracking_number", ""]},
                "shipping_label": {"$ifNull": ["$shipping_label", ""]},
                "shipping_label_preview": {"$ifNull": ["$shipping_label_preview", ""]},
                "shipping_label_print": {"$ifNull": ["$shipping_label_print", ""]},
                "channel_name": {"$ifNull": ["$channel_name", ""]},
                "channel_order_id": {"$ifNull": ["$channel_order_id", ""]},
                "fulfillment_type": {"$ifNull": ["$fulfillment_type", ""]},
                "purchase_order_date": {"$ifNull": ["$purchase_order_date", None]},
                "expected_delivery_date": {"$ifNull": ["$expected_delivery_date", None]},
                "created_at": {"$ifNull": ["$created_at", None]},
                "updated_at": {"$ifNull": ["$updated_at", None]},
                "customer_order_id" : {"$ifNull" : ["$customer_order_id",""]},
                "weight_value" : {"$ifNull" : ['$weight_value',""]},
                "currency" : {"$ifNull" : ['$currency',"USD"]},
                "ordered_products": {
                    "$map": {
                        "input": "$ordered_products",
                        "as": "product",
                        "in": {
                            "product_id": {"$toString": "$$product.product_id"},
                            "product_title": {"$ifNull": ["$$product.title", ""]},
                            "sku": {"$ifNull": ["$$product.sku", ""]},
                            "price": {"$ifNull": ["$$product.unit_price", 0.0]},
                            "quantity": {"$ifNull": ["$$product.quantity", 0]},
                            "quantity_price": {"$ifNull": ["$$product.quantity_price", 0.0]},
                        }
                    }
                }
            }
        }
    ]
    manual_order_details = list(custom_order.objects.aggregate(*pipeline))
    if manual_order_details:
        for i in manual_order_details[0]['ordered_products']:
            pipeline = [
            {"$match": {"_id" : ObjectId(i['product_id'])}},
            {
                "$project": {
                    "_id": None,
                    "image_url": {"$ifNull": ["$image_url", ""]}
                }
            }
            ]
            product_image_obj = list(Product.objects.aggregate(*(pipeline)))
            try:
                i['product_image'] = product_image_obj[0].get('image_url')
            except:
                i['product_image'] = ""
        data['order_details'] = manual_order_details[0]
    else:
        data['error'] = "Manual order not found."
    return data

def ordersCountForDashboard(request):
    from django.utils.timezone import now
    data = dict()
    marketplace_id = request.GET.get('marketplace_id')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    preset = request.GET.get("preset", "Today")
    country = request.GET.get('country', 'US')
    filtered_marketplace_id = get_filtered_marketplaces(country, marketplace_id)
    brand_ids = []
    array_brand_ids = request.GET.getlist('brand_id[]')
    if array_brand_ids:
        brand_ids.extend(array_brand_ids)
    single_brand_id = request.GET.get('brand_id')
    if single_brand_id and single_brand_id not in brand_ids:
        brand_ids.append(single_brand_id)
    product_ids = request.GET.getlist('product_id[]')
    product_id = request.GET.get('product_id', None)
    if product_ids:
        product_ids = [ObjectId(pid) for pid in product_ids]
    elif product_id:
        product_ids = [ObjectId(product_id)]
    else:
        product_ids = None
    timezone_str = "US/Pacific"
    if start_date:
        start_date, end_date = convertdateTotimezone(start_date, end_date, timezone_str)
    else:
        start_date, end_date = get_date_range(preset, timezone_str)
    if timezone_str != 'UTC':
        start_date, end_date = convertLocalTimeToUTC(start_date, end_date, timezone_str)

    match_conditions = {
        "order_date": {"$gte": start_date, "$lte": end_date},
        "order_status": {"$nin": ["Canceled", "Cancelled"]},
        "order_total": {"$gt": 0}
    }
    
    custom_match_conditions = {
        "purchase_order_date": {"$gte": start_date, "$lte": end_date},
        "order_status": {"$nin": ["Canceled", "Cancelled"]},
        "total_price": {"$gt": 0}
    }
    if country:
        match_conditions['geo']=country.upper()
        custom_match_conditions['geo'] = country.upper()

    if brand_ids:
        try:
            brand_object_ids = [ObjectId(bid) for bid in brand_ids if bid]
            if brand_object_ids:
                products = list(Product.objects.filter(brand_id__in=brand_object_ids).only('id'))
                product_ids_from_brands = [p.id for p in products]
                if product_ids_from_brands:
                    if product_ids:
                        product_ids = [pid for pid in product_ids if pid in product_ids_from_brands]
                        if not product_ids:
                            match_conditions["_id"] = None
                            custom_match_conditions["_id"] = None
                    else:
                        product_ids = product_ids_from_brands
                else:
                    match_conditions["_id"] = None
                    custom_match_conditions["_id"] = None
        except Exception as e:
            print(f"Error processing brand filters: {str(e)}")

    def build_pipeline_with_product_filter(base_match_conditions, skip_marketplace_filter=False):
        pipeline = [{"$match": base_match_conditions}]
        
        # Add marketplace filter if filtered_marketplace_id is provided
        if not skip_marketplace_filter and filtered_marketplace_id and isinstance(filtered_marketplace_id, list) and len(filtered_marketplace_id) > 0:
            pipeline[0]["$match"]["marketplace_id"] = {"$in": filtered_marketplace_id}

        if product_ids:
            pipeline.extend([
                {
                    "$lookup": {
                        "from": "order_items",
                        "localField": "order_items",
                        "foreignField": "_id",
                        "as": "order_items_details"
                    }
                },
                {"$unwind": "$order_items_details"},
                {
                    "$match": {
                        "order_items_details.ProductDetails.product_id": {"$in": product_ids}
                    }
                },
                {
                    "$group": {
                        "_id": "$purchase_order_id",
                        "order_total": {"$first": "$order_total"},
                        "order_date": {"$first": "$order_date"},
                        "order_status": {"$first": "$order_status"},
                        "marketplace_id": {"$first": "$marketplace_id"}
                    }
                }
            ])
        return pipeline

    def count_custom_orders(q):
        custom_match = custom_match_conditions.copy()
        # Remove grouping by purchase_order_id to count all orders
        pipeline = [
            {"$match": custom_match},
            {
                "$group": {
                    "_id": None,
                    "count": {"$sum": 1},
                    "order_value": {"$sum": "$total_price"}
                }
            }
        ]
        res = list(custom_order.objects.aggregate(*pipeline))
        if res:
            q.put((res[0].get("count", 0), res[0].get("order_value", 0)))
        else:
            q.put((0, 0))

    if marketplace_id == "all":
        # Get only filtered marketplaces
        filtered_marketplaces = Marketplace.objects.filter(id__in=filtered_marketplace_id).only('id', 'name')
        results = {}
        threads = []
        
        def process_marketplace(mp):
            mp_match_conditions = match_conditions.copy()
            mp_match_conditions["marketplace_id"] = mp.id
            pipeline = build_pipeline_with_product_filter(mp_match_conditions, skip_marketplace_filter=True)
            # Remove grouping by purchase_order_id to count all orders
            pipeline.append({
                "$group": {
                    "_id": None,
                    "count": {"$sum": 1},
                    "order_value": {"$sum": "$order_total"}
                }
            })
            res = list(Order.objects.aggregate(*pipeline))
            count = res[0].get("count", 0) if res else 0
            order_value = round(res[0].get("order_value", 0), 2) if res else 0
            results[mp.name] = {
                "count": count,
                "order_value": order_value
            }
        
        for mp in filtered_marketplaces:
            thread = threading.Thread(target=process_marketplace, args=(mp,))
            thread.start()
            threads.append(thread)
        
        for thread in threads:
            thread.join()
        
        # Calculate total from marketplace results
        total_order_count = sum(mp_data["count"] for mp_data in results.values())
        total_order_value = round(sum(mp_data["order_value"] for mp_data in results.values()), 2)
        
        # Add custom orders to total
        q2 = Queue()
        t2 = threading.Thread(target=count_custom_orders, args=(q2,))
        t2.start()
        t2.join()
        custom_result = q2.get()
        custom_order_count, custom_order_value = custom_result
        
        total_order_count += custom_order_count
        total_order_value = round(total_order_value + custom_order_value, 2)
        
        # Now calculate percentages
        for mp_name, mp_data in results.items():
            percentage = round((mp_data["count"] / total_order_count) * 100, 2) if total_order_count else 0
            data[mp_name] = {
                "count": mp_data["count"],
                "percentage": f"{percentage}%",
                "order_value": mp_data["order_value"]
            }
        
        data['total_order_count'] = {
            "value": total_order_count,
            "percentage": "100.0%",
            "order_value": total_order_value
        }
        
    elif marketplace_id != "all" and marketplace_id != "custom":
        mp_id = ObjectId(marketplace_id)
        mp_match_conditions = match_conditions.copy()
        mp_match_conditions["marketplace_id"] = mp_id
        pipeline = build_pipeline_with_product_filter(mp_match_conditions, skip_marketplace_filter=True)
        # Remove grouping by purchase_order_id to count all orders
        pipeline.append({
            "$group": {
                "_id": None,
                "count": {"$sum": 1},
                "order_value": {"$sum": "$order_total"}
            }
        })
        res = list(Order.objects.aggregate(*pipeline))
        count = res[0].get("count", 0) if res else 0
        order_value = round(res[0].get("order_value", 0), 2) if res else 0
        name = DatabaseModel.get_document(Marketplace.objects, {"id": marketplace_id}, ['name']).name
        data[name] = {
            "value": count,
            "percentage": "100.0%",
            "order_value": order_value
        }
        data['total_order_count'] = {
            "value": count,
            "percentage": "100.0%",
            "order_value": order_value
        }
        
    elif marketplace_id == "custom":
        custom_match = custom_match_conditions.copy()
        # Remove grouping by purchase_order_id to count all orders
        pipeline = [
            {"$match": custom_match},
            {
                "$group": {
                    "_id": None,
                    "count": {"$sum": 1},
                    "order_value": {"$sum": "$total_price"}
                }
            }
        ]
        res = list(custom_order.objects.aggregate(*pipeline))
        count = res[0].get("count", 0) if res else 0
        order_value = round(res[0].get("order_value", 0), 2) if res else 0
        data['custom'] = {
            "value": count,
            "percentage": "100.0%",
            "order_value": order_value
        }
        data['total_order_count'] = {
            "value": count,
            "percentage": "100.0%",
            "order_value": order_value
        }

    return sanitize_data(data)

def totalSalesAmount(request):
    data = dict()
    marketplace_id = request.GET.get('marketplace_id')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    preset = request.GET.get("preset", "Today")
    timezone_str = "US/Pacific"
    brand_id_list = request.GET.get('brand_id')
    product_id = request.GET.get("product_id")
    manufacturer_name = request.GET.get("manufacturer_name")
    fulfillment_channel = request.GET.get("fulfillment_channel")
    if product_id:
        try:
            product_id = json.loads(product_id)
        except Exception:
            product_id = [product_id]
    if brand_id_list:
        try:
            brand_id_list = json.loads(brand_id_list)
        except Exception:
            brand_id_list = [brand_id_list]
    if manufacturer_name:
        try:
            manufacturer_name = json.loads(manufacturer_name)
        except Exception:
            manufacturer_name = [manufacturer_name]
    if start_date:
        start_date, end_date = convertdateTotimezone(start_date, end_date, timezone_str)
    else:
        start_date, end_date = get_date_range(preset, timezone_str)
    if timezone_str != 'UTC':
        start_date, end_date = convertLocalTimeToUTC(start_date, end_date, timezone_str)
    orders = grossRevenue(
        start_date, end_date,
        marketplace_id=marketplace_id,
        brand_id=brand_id_list,
        product_id=product_id,
        manufacuture_name=manufacturer_name,
        fulfillment_channel=fulfillment_channel,
        timezone=timezone_str
    )
    total_sales = sum(order['order_total'] for order in orders)
    data['total_sales'] = round(total_sales, 2)
    return data

@csrf_exempt
def salesAnalytics(request):
    try:
        data = dict()
        json_request = JSONParser().parse(request)
        marketplace_id = json_request.get('marketplace_id', 'all')  
        country=json_request.get('country','US')
        filtered_marketplace_id=get_filtered_marketplaces(country,marketplace_id)
        start_date = json_request.get('start_date')  
        end_date = json_request.get('end_date')  
        timezone_str = json_request.get('timezone', 'US/Pacific')
        brand_id_list = json_request.get('brand_id')
        preset = json_request.get("preset", "Today")     
        product_id_list = json_request.get('product_id')
        if start_date and start_date != "":
            if isinstance(start_date, datetime):
                start_date = start_date.isoformat()
            if isinstance(end_date, datetime):
                end_date = end_date.isoformat()
            start_date, end_date = convertdateTotimezone(start_date, end_date, timezone_str)
        else:
            start_date, end_date = get_date_range(preset, timezone_str)
        orders = grossRevenue(
            start_date, end_date, 
            filtered_marketplace_id=filtered_marketplace_id, 
            brand_id=brand_id_list, 
            timezone=timezone_str,
            product_id=product_id_list
        )
        orders = [o for o in orders if o.get('order_status') not in ['Cancelled','Canceled'] and o.get('order_total', 0) > 0]
        data['total_sales'] = sum(o['original_order_total'] for o in orders)
        order_days = defaultdict(lambda: {'order_count': 0, 'order_value': 0.0})
        for o in orders:
            order_date = o['order_date']
            if isinstance(order_date, str):
                order_date = datetime.fromisoformat(order_date)
            date_key = order_date.strftime('%Y-%m-%d')
            order_days[date_key]['order_count'] += 1
            order_days[date_key]['order_value'] += o['order_total']
        data['order_days'] = [
            {'date': k, 'order_count': v['order_count'], 'order_value': v['order_value']}
            for k, v in sorted(order_days.items())
        ]
        sanitized_data = sanitize_floats(data)
        return sanitized_data
    except Exception as e:
        return {"error": str(e)}
    
@csrf_exempt
def mostSellingProducts(request):
    data = dict()
    pipeline = list()
    marketPlaceId = request.GET.get('marketPlaceId')
    start_date = request.GET.get('start_date')  
    end_date = request.GET.get('end_date')  
    match_conditions = {}
    if start_date and end_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d')
        end_date = datetime.strptime(end_date, '%Y-%m-%d')
        match_conditions["order_data.order_date"] = {"$gte": start_date, "$lte": end_date}
    pipeline.extend([
        {
            "$lookup": {
                "from": "order",
                "localField": "_id",
                "foreignField": "order_items",
                "as": "order_data"
            }
        },
        {
            "$unwind": "$order_data"
        },
        {
            "$match": match_conditions
        },
        {
            "$lookup": {
                "from": "product",
                "localField": "ProductDetails.product_id",
                "foreignField": "_id",
                "as": "product_ins"
            }
        },
        {
            "$unwind": "$product_ins"
        },
        {
            "$lookup": {
                "from": "marketplace",
                "localField": "product_ins.marketplace_id",
                "foreignField": "_id",
                "as": "marketplace_ins"
            }
        },
        {
            "$unwind": "$marketplace_ins"
        },
    ])
    if marketPlaceId and marketPlaceId != "all" and marketPlaceId != "custom":
        match = {
            "$match": {
                "product_ins.marketplace_id": ObjectId(marketPlaceId)
            }
        }
        pipeline.append(match)
    pipeline2 = [
        {
            "$group": {
                "_id": {
                    "product_id": "$ProductDetails.product_id",
                    "product_title": "$product_ins.product_title",
                    "sku": "$product_ins.sku",
                    "image_url": "$product_ins.image_url",
                    "price": "$product_ins.price",
                    "channel_name" : "$marketplace_ins.name"
                },
                "total_quantity_sold": {"$sum": "$ProductDetails.QuantityOrdered"},
                "total_revenue": {"$sum": {"$multiply": ["$ProductDetails.QuantityOrdered", "$product_ins.price"]}}
            }
        },
        {
            "$project": {
                "_id": 0,
                "product_id": {"$toString": "$_id.product_id"},
                "product_title": "$_id.product_title",
                "sku": "$_id.sku",
                "product_image": "$_id.image_url",
                "price": "$_id.price",
                "channel_name" : "$_id.channel_name",
                "sales_count": "$total_quantity_sold",
                "revenue": {"$round": ["$total_revenue", 2]}
            }
        },
        {
            "$sort": {"sales_count": -1}
        },
        {
            "$limit": 7
        },
    ]
    pipeline.extend(pipeline2)
    top_products = list(OrderItems.objects.aggregate(*pipeline))
    custom_match_conditions = {}
    if start_date and end_date:
        custom_match_conditions["purchase_order_date"] = {"$gte": start_date, "$lte": end_date}
    custom_pipeline = [
        {
            "$match": custom_match_conditions
        },
        {
            "$unwind": "$ordered_products"
        },
        {
            "$group": {
                "_id": {
                    "product_id": "$ordered_products.product_id",
                    "product_title": "$ordered_products.title",
                    "sku": "$ordered_products.sku",
                    "price": "$ordered_products.unit_price"
                },
                "total_quantity_sold": {"$sum": "$ordered_products.quantity"},
                "total_revenue": {"$sum": "$ordered_products.quantity_price"}
            }
        },
        {
            "$project": {
                "_id": 0,
                "product_id": {"$toString": "$_id.product_id"},
                "product_title": "$_id.product_title",
                "sku": "$_id.sku",
                "price": "$_id.price",
                "sales_count": "$total_quantity_sold",
                "revenue": {"$round": ["$total_revenue", 2]}
            }
        },
        {
            "$sort": {"sales_count": -1}
        },
        {
            "$limit": 7
        }
    ]
    custom_top_products = list(custom_order.objects.aggregate(*custom_pipeline))
    for i in custom_top_products:
        pipeline = [
        {"$match": {"_id" : ObjectId(i['product_id'])}},
        {
            "$project": {
                "_id": None,
                "image_url": {"$ifNull": ["$image_url", ""]}
            }
        }
        ]
        product_image_obj = list(Product.objects.aggregate(*(pipeline)))
        try:
            i['product_image'] = product_image_obj[0].get('image_url')
        except:
            print(i['product_id'])
            i['product_image'] = ""
        i['channel_name'] = "custom"
    if marketPlaceId == "custom":
        data['top_products'] = custom_top_products
        return data
    if marketPlaceId == "all":
        top_products.extend(custom_top_products)
        top_products = sorted(top_products, key=lambda x: x['sales_count'], reverse=True)[:7]
    data['top_products'] = top_products
    return data
def change_sign(value):
    """
    Change the sign of a given value only if it's negative.
    Args:
    value (number): The input value to potentially change the sign of.
    Returns:
    number: The input value with its sign changed if it was negative, otherwise unchanged.
    """
    if value < 0:
        return -value
    else:
        return value
@csrf_exempt
def getSalesTrendPercentage(request):
    from pytz import timezone
    from rest_framework.parsers import JSONParser
    data = dict()
    json_request = JSONParser().parse(request)
    range_type = json_request.get('range_type', 'month')  
    marketplace_id = json_request.get('marketplace_id')  
    timezone_str = json_request.get('timezone', 'US/Pacific')
    local_tz = timezone(timezone_str)
    now = datetime.now(local_tz)
    if now.tzinfo is None:
        now = local_tz.localize(now)
    now = now.astimezone(pytz.UTC)
    now = now.replace(tzinfo=None)
    if range_type == 'day':
        current_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        previous_start = current_start - timedelta(days=1)
        previous_end = current_start
    elif range_type == 'week':
        current_start = now - timedelta(days=now.weekday())
        current_start = current_start.replace(hour=0, minute=0, second=0, microsecond=0)
        previous_start = current_start - timedelta(weeks=1)
        previous_end = current_start
    elif range_type == 'month':
        current_start = datetime(now.year, now.month, 1)
        previous_month_end = current_start - timedelta(days=1)
        previous_start = datetime(previous_month_end.year, previous_month_end.month, 1)
        previous_end = current_start
    elif range_type == 'year':
        current_start = datetime(now.year, 1, 1)
        previous_start = datetime(now.year - 1, 1, 1)
        previous_end = current_start
    else:
        data['error'] = "Invalid range_type provided."
        return data
    match_pipeline = [
        {
            "$facet": {
                "current_range": [
                    {
                        "$match": {
                            "order_date": {
                                "$gte": current_start,
                                "$lt": now
                            },
                            **({"marketplace_id": ObjectId(marketplace_id)} if marketplace_id and marketplace_id != "custom" and marketplace_id != "all" else {})
                        }
                    },
                    {
                        "$group": {
                            "_id": "$marketplace_id",
                            "sales_value": {"$sum": "$order_total"}
                        }
                    }
                ],
                "previous_range": [
                    {
                        "$match": {
                            "order_date": {
                                "$gte": previous_start,
                                "$lt": previous_end
                            },
                            **({"marketplace_id": ObjectId(marketplace_id)} if marketplace_id and marketplace_id != "custom" and marketplace_id != "all" else {})
                        }
                    },
                    {
                        "$group": {
                            "_id": "$marketplace_id",
                            "sales_value": {"$sum": "$order_total"}
                        }
                    }
                ]
            }
        }
    ]
    custom_match_pipeline = [
        {
            "$facet": {
                "current_range": [
                    {
                        "$match": {
                            "purchase_order_date": {
                                "$gte": current_start,
                                "$lt": now
                            }
                        }
                    },
                    {
                        "$group": {
                            "_id": None,
                            "sales_value": {"$sum": "$total_price"}
                        }
                    }
                ],
                "previous_range": [
                    {
                        "$match": {
                            "purchase_order_date": {
                                "$gte": previous_start,
                                "$lt": previous_end
                            }
                        }
                    },
                    {
                        "$group": {
                            "_id": None,
                            "sales_value": {"$sum": "$total_price"}
                        }
                    }
                ]
            }
        }
    ]
    if marketplace_id == "custom":
        trend_data = list(custom_order.objects.aggregate(*custom_match_pipeline))
    else:
        trend_data = list(Order.objects.aggregate(*match_pipeline))
        custom_trend_data = list(custom_order.objects.aggregate(*custom_match_pipeline))
        if custom_trend_data:
            for key in ["current_range", "previous_range"]:
                for item in custom_trend_data[0][key]:
                    trend_data[0][key].append({"_id": "custom", "sales_value": item["sales_value"]})
    if trend_data:
        current_range_data = {item["_id"]: sanitize_data(item["sales_value"]) for item in trend_data[0]["current_range"]}
        previous_range_data = {item["_id"]: sanitize_data(item["sales_value"]) for item in trend_data[0]["previous_range"]}
        if marketplace_id == "all":  
            current_total = sum(current_range_data.values())
            previous_total = sum(previous_range_data.values())
            percentage_change = ((current_total - previous_total) / previous_total * 100) if previous_total != 0 else (100 if current_total > 0 else 0)
            current_percentage = (current_total / previous_total * 100) if previous_total != 0 else (100 if current_total > 0 else 0)
            data['trend_percentage'] = [{
                "id": "All Channels",
                "current_range_sales": current_total,
                "previous_range_sales": previous_total,
                "trend_percentage": round(percentage_change, 2),
                "current_percentage": round(current_percentage, 2)
            }]
        elif marketplace_id == "custom":  
            current_total = current_range_data.get(None, 0)
            previous_total = previous_range_data.get(None, 0)
            percentage_change = ((current_total - previous_total) / previous_total * 100) if previous_total != 0 else (100 if current_total > 0 else 0)
            current_percentage = (current_total / previous_total * 100) if previous_total != 0 else (100 if current_total > 0 else 0)
            data['trend_percentage'] = [{
                "id": "Custom Orders",
                "current_range_sales": current_total,
                "previous_range_sales": previous_total,
                "trend_percentage": round(percentage_change, 2),
                "current_percentage": round(current_percentage, 2)
            }]
        else:  
            trend_percentage = []
            marketplace_name = DatabaseModel.get_document(Marketplace.objects, {"id": marketplace_id}, ['name']).name
            for key in set(current_range_data.keys()).union(previous_range_data.keys()):
                current_value = current_range_data.get(key, 0)
                previous_value = previous_range_data.get(key, 0)
                percentage_change = ((current_value - previous_value) / previous_value * 100) if previous_value != 0 else (100 if current_value > 0 else 0)
                current_percentage = (current_value / previous_value * 100) if previous_value != 0 else (100 if current_value > 0 else 0)
                trend_percentage.append({
                    "id": str(marketplace_name),
                    "current_range_sales": current_value,
                    "previous_range_sales": previous_value,
                    "trend_percentage": round(percentage_change, 2),
                    "current_percentage": round(current_percentage, 2)
                })
            data['trend_percentage'] = trend_percentage
    else:
        data['trend_percentage'] = []
    return data
@csrf_exempt
def fetchSalesSummary(request):
    data = {}
    total_sales_pipeline = []
    pipeline = []
    match = {}
    json_request = JSONParser().parse(request)
    marketplace_id = json_request.get('marketplace_id')
    start_date = json_request.get('start_date')  
    end_date = json_request.get('end_date')  
    if start_date and end_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d')
        end_date = datetime.strptime(end_date, '%Y-%m-%d')
        match["order_date"] = {"$gte": start_date, "$lte": end_date}
    if marketplace_id != None and marketplace_id != "all" and marketplace_id != "" and marketplace_id != "custom":
        match["marketplace_id"] = ObjectId(marketplace_id)
    if marketplace_id == "all" or marketplace_id != "custom":
        if match:
            match_stage = {"$match": match}
            total_sales_pipeline.append(match_stage)
            pipeline.append(match_stage)
        total_sales_pipeline.extend([
            {
                "$group": {
                    "_id": None,  
                    "total_sales": {"$sum": "$order_total"}
                }
            },
            {
                "$project": {
                    "_id": 0,
                    "total_sales": {"$round": ["$total_sales", 2]}
                }
            }
        ])
        summary = list(Order.objects.aggregate(*total_sales_pipeline))
        if summary:
            data['total_sales'] = summary[0].get('total_sales', 0)
        else:
            data['total_sales'] = 0
        pipeline.extend([
            {
                "$unwind": "$order_items"  
            },
            {
                "$group": {
                    "_id": None,  
                    "ids": {"$addToSet": "$order_items"}  
                }
            },
            {
                "$project": {
                    "_id": 0,
                    "ids": 1
                }
            },
        ])
        summary1 = list(Order.objects.aggregate(*pipeline))
        if summary1:
            s_pipeline = [
                {"$match": {
                    "_id": {"$in": summary1[0]['ids']}
                }},
                {
                    "$group": {
                        "_id": None,
                        "total_units_sold": {"$sum": "$ProductDetails.QuantityOrdered"},
                        "unique_product_ids": {"$addToSet": "$ProductDetails.product_id"},
                    }
                },
                {
                    "$project": {
                        "_id": 0,
                        "total_units_sold": 1,
                        "total_sold_product_count": {"$size": "$unique_product_ids"},
                    }
                }
            ]
            summary2 = list(OrderItems.objects.aggregate(*s_pipeline))
            if summary2:
                data['total_units_sold'] = summary2[0].get('total_units_sold', 0)
                data['total_sold_product_count'] = summary2[0].get('total_sold_product_count', 0)
        else:
            data['total_units_sold'] = 0
            data['total_sold_product_count'] = 0
    custom_match = {}
    if start_date and end_date:
        custom_match["purchase_order_date"] = {"$gte": start_date, "$lte": end_date}
    custom_pipeline = [
        {"$match": custom_match},
        {"$unwind": "$ordered_products"},
        {
            "$group": {
                "_id": None,
                "total_custom_sales": {"$sum": "$ordered_products.quantity_price"},
                "total_custom_units_sold": {"$sum": "$ordered_products.quantity"},
                "unique_custom_product_ids": {"$addToSet": "$ordered_products.product_id"}
            }
        },
        {
            "$project": {
                "_id": 0,
                "total_custom_sales": {"$round": ["$total_custom_sales", 2]},
                "total_custom_units_sold": 1,
                "total_custom_sold_product_count": {"$size": "$unique_custom_product_ids"}
            }
        }
    ]
    custom_summary = list(custom_order.objects.aggregate(*custom_pipeline))
    if custom_summary:
        total_custom_sales = custom_summary[0].get('total_custom_sales', 0)
        total_custom_units_sold = custom_summary[0].get('total_custom_units_sold', 0)
        total_custom_sold_product_count = custom_summary[0].get('total_custom_sold_product_count', 0)
    else:
        total_custom_sales = 0
        total_custom_units_sold = 0
        total_custom_sold_product_count = 0
    if marketplace_id == "custom":
        data['total_sales'] = total_custom_sales
        data['total_units_sold'] = total_custom_units_sold
        data['total_sold_product_count'] = total_custom_sold_product_count
    if marketplace_id == "": 
        data['total_sales'] += total_custom_sales
        data['total_units_sold'] += total_custom_units_sold
        data['total_sold_product_count'] += total_custom_sold_product_count
    return data
@csrf_exempt
def fetchTopSellingCategories(request):
    data = dict()
    json_request = JSONParser().parse(request)
    marketplace_id = json_request.get('marketplace_id')  
    limit = int(json_request.get('limit', 15))  
    start_date = json_request.get('start_date')  
    end_date = json_request.get('end_date')  
    match_conditions = {}
    if start_date and end_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d')
        end_date = datetime.strptime(end_date, '%Y-%m-%d')
        match_conditions["order_data.order_date"] = {"$gte": start_date, "$lte": end_date}
    if marketplace_id != None and marketplace_id != "all" and marketplace_id != "custom":
        match_conditions["order_data.marketplace_id"] = ObjectId(marketplace_id)
    if marketplace_id != "custom":
        pipeline = [
            {
                "$lookup": {
                    "from": "order",
                    "localField": "_id",
                    "foreignField": "order_items",
                    "as": "order_data"
                }
            },
            {
                "$unwind": "$order_data"
            },
            {
                "$match": match_conditions
            },
            {
                "$lookup": {
                    "from": "product",
                    "localField": "ProductDetails.product_id",
                    "foreignField": "_id",
                    "as": "product_ins"
                }
            },
            {
                "$unwind": "$product_ins"
            },
            {
                "$lookup": {
                    "from": "category",
                    "localField": "product_ins.category",
                    "foreignField": "name",
                    "as": "category_ins"
                }
            },
            {
                "$unwind": "$category_ins"
            },
            {
                "$group": {
                    "_id": {
                        "category_id": "$category_ins._id",
                        "category_name": "$category_ins.name"
                    },
                    "total_order_value": {"$sum": {"$multiply": ["$ProductDetails.QuantityOrdered", "$product_ins.price"]}}
                }
            },
            {
                "$project": {
                    "_id": 0,
                    "category_id": {"$toString": "$_id.category_id"},
                    "category_name": "$_id.category_name",
                    "total_order_value": {"$round": ["$total_order_value", 2]}
                }
            },
            {
                "$sort": {"total_order_value": -1}
            },
            {
                "$limit": limit
            }
        ]
        top_categories = list(OrderItems.objects.aggregate(*pipeline))
    custom_match_conditions = {}
    if start_date and end_date:
        custom_match_conditions["purchase_order_date"] = {"$gte": start_date, "$lte": end_date}
    custom_pipeline = [
        {"$match": custom_match_conditions},
        {"$unwind": "$ordered_products"},
        {
            "$lookup": {
                "from": "product",
                "localField": "ordered_products.product_id",
                "foreignField": "_id",
                "as": "product_ins"
            }
        },
        {
            "$unwind": "$product_ins"
        },
        {
            "$lookup": {
                "from": "category",
                "localField": "product_ins.category",
                "foreignField": "name",
                "as": "category_ins"
            }
        },
        {
            "$unwind": "$category_ins"
        },
        {
            "$group": {
                "_id": {
                    "category_id": "$category_ins._id",
                    "category_name": "$category_ins.name"
                },
                "total_order_value": {"$sum": "$ordered_products.quantity_price"}
            }
        },
        {
            "$project": {
                "_id": 0,
                "category_id": {"$toString": "$_id.category_id"},
                "category_name": "$_id.category_name",
                "total_order_value": {"$round": ["$total_order_value", 2]}
            }
        },
        {
            "$sort": {"total_order_value": -1}
        },
        {
            "$limit": limit
        }
    ]
    custom_top_categories = list(custom_order.objects.aggregate(*custom_pipeline))
    if marketplace_id == "custom":
        data['top_categories'] = custom_top_categories
    else:
        combined_categories = {}
        for category in top_categories + custom_top_categories:
            category_id = category["category_id"]
            if category_id in combined_categories:
                combined_categories[category_id]["total_order_value"] += category["total_order_value"]
            else:
                combined_categories[category_id] = category
        sorted_categories = sorted(combined_categories.values(), key=lambda x: x["total_order_value"], reverse=True)[:limit]
        data['top_categories'] = sorted_categories
    return data
@csrf_exempt
def createUser(request):
    data = dict()
    json_request = JSONParser().parse(request)
    email = json_request.get("email")
    old_user_obj = DatabaseModel.get_document(user.objects,{"email" : email},['id'])
    if old_user_obj == None:
        user_data = {
            "first_name": json_request.get("first_name"),
            "last_name" : json_request.get('last_name'),
            "email": email,
            "password": json_request.get("password"),  
            "role_id": ObjectId(json_request.get("role_id")),
            "profile_image" : json_request.get("profile_image", "")
        }
        new_user = DatabaseModel.save_documents(user, user_data)
        data["message"] = "User created successfully."
        data["user_id"] = str(new_user.id)
    else:
        data["message"] = "User Already Present."
        data["user_id"] = str(old_user_obj.id)
    return data
@csrf_exempt
def updateUser(request):
    data = dict()
    json_request = JSONParser().parse(request)
    target_user_id = json_request.get("target_user_id")
    update_obj = json_request.get('update_obj')
    old_user_obj = DatabaseModel.get_document(user.objects,{"id" : target_user_id})
    data["message"] = "User Not Updated."
    try:
        update_obj['role_id'] = ObjectId(update_obj['role_id'])
    except:
        pass
    if old_user_obj:
        DatabaseModel.update_documents(user.objects,{"id" : target_user_id},update_obj)
        data["message"] = "User updated successfully."
    return data
@csrf_exempt
def listUsers(request):
    data = dict()
    limit = int(request.GET.get("limit", 100))  
    skip = int(request.GET.get("skip", 0))  
    pipeline = [
         {
            "$lookup": {
                "from": "role",
                "localField": "role_id",
                "foreignField": "_id",
                "as": "role_ins"
            }
        },
        {
            "$unwind": {
                "path": "$role_ins",
                "preserveNullAndEmptyArrays": True  
            }
        },
        {
        "$project": {
            "_id": 0,
            "id": {"$toString": "$_id"},
            "first_name": 1,
            "last_name": 1,
            "email" : 1,
            "role_name": "$role_ins.name",
            "creation_date" :1,
            "role_id" : {"$toString" : "$role_id"}
        }
    },
        {"$skip": skip},
        {"$limit": limit},
        {
            "$sort" : {
                "id" : -1
            }
        }
    ]
    users = list(user.objects.aggregate(*pipeline))
    data["users"] = users
    return data
def fetchUserDetails(request):
    data = dict()
    target_user_id = request.GET.get("target_user_id")
    data['user_obj'] = {}
    pipeline = [
    {
        "$match" : {
            "_id" : ObjectId(target_user_id)
        }
    },
    {
        "$lookup": {
            "from": "role",
            "localField": "role_id",
            "foreignField": "_id",
            "as": "role_ins"
        }
    },
    {
        "$unwind": "$role_ins"
    },
    {
        "$project": {
            "_id": 0,
            "id": {"$toString": "$_id"},
            "first_name": 1,
            "last_name": 1,
            "email" : 1,
            "mobile_number" : 1,
            "profile_image" : 1,
            "role_name": "$role_ins.name",
        }
    },
    ]
    user_obj = list(user.objects.aggregate(*pipeline))
    if user_obj != []:
        data['user_obj'] = user_obj[0]
    return data
def fetchRoles(request):
    pipeline = [
        {
            "$project": {
                "_id": 0,
                "id": {"$toString": "$_id"},
                "name": 1,
            }
        }
    ]
    role_list = list(role.objects.aggregate(*pipeline))
    return role_list
@csrf_exempt
def fetchInventryList(request):
    data = dict()
    json_request = JSONParser().parse(request)
    marketplace_id = json_request.get('marketplace_id')
    skip = int(json_request.get('skip'))
    limit = int(json_request.get('limit'))
    search_query = json_request.get('search_query')   
    sort_by = json_request.get('sort_by')
    sort_by_value = json_request.get('sort_by_value')
    pipeline = []
    count_pipeline = []
    match = {}
    if marketplace_id != None and marketplace_id != "" and marketplace_id != "all":
        match['marketplace_id'] = ObjectId(marketplace_id)
    if search_query != None and search_query != "":
        search_query = re.escape(search_query.strip())
        match["$or"] = [
            {"product_title": {"$regex": search_query, "$options": "i"}},
            {"sku": {"$regex": search_query, "$options": "i"}},
        ]
    if match != {}:
        match_pipeline = {
            "$match" : match}
        print(match_pipeline)
        pipeline.append(match_pipeline)
        count_pipeline.append(match_pipeline)
    pipeline.extend([
        {
            "$lookup" : {
                "from" : "marketplace",
                "localField" : "marketplace_id",
                "foreignField" : "_id",
                "as" : "marketplace"
            }
        },
        {
            "$unwind" : "$marketplace"
        },
        {
            "$project" : {
                "_id" : 0,
                "id" : {"$toString" : "$_id"},
                "product_title" : 1,
                "sku" : 1,
                "price" : 1,
                "quantity" : 1,
                "image_url" : {"$ifNull" : ["$image_url",""]},  
                "marketplace_name" : "$marketplace.name",
            }
        },
        {
            "$skip" : skip
        },
        {
            "$limit" : limit+skip
        }
    ])
    if sort_by != None and sort_by != "":
        sort = {
            "$sort" : {
                sort_by : int(sort_by_value)
            }
        }
        pipeline.append(sort)
    inventry_list = list(Product.objects.aggregate(*(pipeline)))
    count_pipeline.extend([
        {
            "$count": "total_count"
        }
    ])
    total_count_result = list(Product.objects.aggregate(*(count_pipeline)))
    total_count = total_count_result[0]['total_count'] if total_count_result else 0
    data['total_count'] = total_count
    data['inventry_list'] = inventry_list
    return data
def exportOrderReport(request):
    orders = list(Order.objects.all())
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Order Report"
    max_products = max(len(order.order_items) for order in orders) if orders else 1
    order_headers = [
        "Purchase Order Id", "Customer Order Id", "Order Date", "Marketplace Name", "Earliest Ship Date",
        "Fulfilment Channel", "Order Status", "Ship Service Level", "Customer Email Id",
        "Has Regulated Items", "Is Replacement Order", "Shipping Information"
    ]
    product_headers = []
    for i in range(1, max_products + 1):
        product_headers.extend([
            f"Product {i} Name", f"Product {i} SKU", f"Product {i} Quantity Ordered",
            f"Product {i} Quantity Shipped", f"Product {i} Unit Price"
        ])
    fixed_headers = ["Discount", "Tax", "Total Order", "Currency"]
    headers = order_headers + product_headers + fixed_headers
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    bold_font = Font(bold=True)
    for col_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_index, value=header)
        cell.fill = blue_fill
        cell.font = bold_font
    row = 2
    for order in orders:
        base_col = 1
        if order.shipping_information:
            shipping_info = order.shipping_information
            postal_address = shipping_info.get('postalAddress', {})
            shipping_details = f"Name: {postal_address.get('name', '')}, " \
                               f"Address1: {postal_address.get('address1', '')}, " \
                               f"Address2: {postal_address.get('address2', '')}, " \
                               f"City: {postal_address.get('city', '')}, " \
                               f"State: {postal_address.get('state', '')}, " \
                               f"PostalCode: {postal_address.get('postalCode', '')}, " \
                               f"Country: {postal_address.get('country', '')}, " \
                               f"Phone: {shipping_info.get('phone', '')}, " \
                               f"MethodCode: {shipping_info.get('methodCode', '')}, " \
                               f"EstimatedShipDate: {shipping_info.get('estimatedShipDate', '')}, " \
                               f"EstimatedDeliveryDate: {shipping_info.get('estimatedDeliveryDate', '')}"
        else:
            shipping_details = None
        order_data = [
            order.purchase_order_id, order.customer_order_id,
            order.order_date.strftime('%Y-%m-%d') if order.order_date else None,
            order.marketplace_id.name if order.marketplace_id else None,
            order.earliest_ship_date, order.fulfillment_channel,
            order.order_status, order.ship_service_level,
            order.customer_email_id, order.has_regulated_items,
            order.is_replacement_order, shipping_details
        ]
        for col_index, value in enumerate(order_data, start=base_col):
            sheet.cell(row=row, column=col_index, value=value)
        col = base_col + len(order_data)
        for product_index, product in enumerate(order.order_items[:max_products]):
            product_details = [
                product.ProductDetails.Title, product.ProductDetails.SKU,
                product.ProductDetails.QuantityOrdered, product.ProductDetails.QuantityShipped,
                product.Pricing['ItemPrice']['Amount'] if product.Pricing else None
            ]
            for col_index, value in enumerate(product_details, start=col + product_index * 5):
                sheet.cell(row=row, column=col_index, value=value)
        col = base_col + len(order_data) + (max_products * 5)
        fixed_details = [
            order.discount if hasattr(order, 'discount') else None,
            order.tax if hasattr(order, 'tax') else None,
            order.order_total, order.currency
        ]
        for col_index, value in enumerate(fixed_details, start=col):
            sheet.cell(row=row, column=col_index, value=value)
        row += 1
    for col in range(1, sheet.max_column + 1):
        sheet.column_dimensions[get_column_letter(col)].width = 20
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    response = HttpResponse(output, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="Order_Report.xlsx"'
    return response
@csrf_exempt
def fetchSalesSummary(request):
    data = {}
    total_sales_pipeline = []
    pipeline = []
    match = {}
    json_request = JSONParser().parse(request)
    marketplace_id = json_request.get('marketplace_id')
    start_date = json_request.get('start_date')  
    end_date = json_request.get('end_date')  
    if start_date and end_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d')
        end_date = datetime.strptime(end_date, '%Y-%m-%d')
        match["order_date"] = {"$gte": start_date, "$lte": end_date}
    if marketplace_id != None and marketplace_id != "all" and marketplace_id != "" and marketplace_id != "custom":
        match["marketplace_id"] = ObjectId(marketplace_id)
    if marketplace_id == "all" or marketplace_id != "custom":
        if match:
            match_stage = {"$match": match}
            total_sales_pipeline.append(match_stage)
            pipeline.append(match_stage)
        total_sales_pipeline.extend([
            {
                "$group": {
                    "_id": None,  
                    "total_sales": {"$sum": "$order_total"},
                    "total_cogs": {"$sum": "$cogs"},  
                    "total_refunds": {"$sum": "$refunds"}  
                }
            },
            {
                "$project": {
                    "_id": 0,
                    "total_sales": {"$round": ["$total_sales", 2]},
                    "total_cogs": {"$round": ["$total_cogs", 2]},
                    "total_refunds": {"$round": ["$total_refunds", 2]}
                }
            }
        ])
        summary = list(Order.objects.aggregate(*total_sales_pipeline))
        if summary:
            data['total_sales'] = summary[0].get('total_sales', 0)
            data['total_cogs'] = summary[0].get('total_cogs', 0)
            data['total_refunds'] = summary[0].get('total_refunds', 0)
            data['margin'] = ((data['total_sales'] - data['total_cogs']) / data['total_sales']) * 100 if data['total_sales'] else 0
        else:
            data['total_sales'] = 0
            data['total_cogs'] = 0
            data['total_refunds'] = 0
            data['margin'] = 0
        pipeline.extend([
            {
                "$unwind": "$order_items"  
            },
            {
                "$group": {
                    "_id": None,  
                    "ids": {"$addToSet": "$order_items"}  
                }
            },
            {
                "$project": {
                    "_id": 0,
                    "ids": 1
                }
            },
        ])
        summary1 = list(Order.objects.aggregate(*pipeline))
        if summary1:
            s_pipeline = [
                {"$match": {
                    "_id": {"$in": summary1[0]['ids']}
                }},
                {
                    "$group": {
                        "_id": None,
                        "total_units_sold": {"$sum": "$ProductDetails.QuantityOrdered"},
                        "unique_product_ids": {"$addToSet": "$ProductDetails.product_id"},
                    }
                },
                {
                    "$project": {
                        "_id": 0,
                        "total_units_sold": 1,
                        "total_sold_product_count": {"$size": "$unique_product_ids"},
                    }
                }
            ]
            summary2 = list(OrderItems.objects.aggregate(*s_pipeline))
            if summary2:
                data['total_units_sold'] = summary2[0].get('total_units_sold', 0)
                data['total_sold_product_count'] = summary2[0].get('total_sold_product_count', 0)
        else:
            data['total_units_sold'] = 0
            data['total_sold_product_count'] = 0
    custom_match = {}
    if start_date and end_date:
        custom_match["purchase_order_date"] = {"$gte": start_date, "$lte": end_date}
    custom_pipeline = [
        {"$match": custom_match},
        {"$unwind": "$ordered_products"},
        {
            "$group": {
                "_id": None,
                "total_custom_sales": {"$sum": "$ordered_products.quantity_price"},
                "total_custom_cogs": {"$sum": "$ordered_products.cogs"},  
                "total_custom_units_sold": {"$sum": "$ordered_products.quantity"},
                "unique_custom_product_ids": {"$addToSet": "$ordered_products.product_id"}
            }
        },
        {
            "$project": {
                "_id": 0,
                "total_custom_sales": {"$round": ["$total_custom_sales", 2]},
                "total_custom_cogs": {"$round": ["$total_custom_cogs", 2]},
                "total_custom_units_sold": 1,
                "total_custom_sold_product_count": {"$size": "$unique_custom_product_ids"}
            }
        }
    ]
    custom_summary = list(custom_order.objects.aggregate(*custom_pipeline))
    if custom_summary:
        total_custom_sales = custom_summary[0].get('total_custom_sales', 0)
        total_custom_cogs = custom_summary[0].get('total_custom_cogs', 0)
        total_custom_units_sold = custom_summary[0].get('total_custom_units_sold', 0)
        total_custom_sold_product_count = custom_summary[0].get('total_custom_sold_product_count', 0)
    else:
        total_custom_sales = 0
        total_custom_cogs = 0
        total_custom_units_sold = 0
        total_custom_sold_product_count = 0
    if marketplace_id == "custom":
        data['total_sales'] = total_custom_sales
        data['total_cogs'] = total_custom_cogs
        data['total_units_sold'] = total_custom_units_sold
        data['total_sold_product_count'] = total_custom_sold_product_count
        data['margin'] = ((total_custom_sales - total_custom_cogs) / total_custom_sales) * 100 if total_custom_sales else 0
    if marketplace_id == "": 
        data['total_sales'] += total_custom_sales
        data['total_cogs'] += total_custom_cogs
        data['total_units_sold'] += total_custom_units_sold
        data['total_sold_product_count'] += total_custom_sold_product_count
        data['margin'] = ((data['total_sales'] - data['total_cogs']) / data['total_sales']) * 100 if data['total_sales'] else 0
    return data
def getProductVariant(request):
    variant_list = list()
    product_id = request.GET.get('product_id')
    is_duplicate = request.GET.get('is_duplicate',False)
    start_date=request.GET.get('start_date',None)
    end_date=request.GET.get('end_date',None)
    preset=request.GET.get('preset',"Today")
    timezone_str='US/Pacific'
    product=DatabaseModel.get_document(Product.objects,{"id":product_id},['parent_sku'])
    if not product or not product.parent_sku:
        return []
    parent_sku = product.parent_sku
    if start_date and start_date != "":
        start_date, end_date = convertdateTotimezone(start_date, end_date, timezone_str)
    else:
        start_date, end_date = get_date_range(preset, timezone_str)
    today_start_date, today_end_date = get_date_range("Today", timezone_str)
    if timezone_str != 'UTC':
        today_start_date, today_end_date = convertLocalTimeToUTC(today_start_date, today_end_date, timezone_str)
        start_date, end_date = convertLocalTimeToUTC(start_date, end_date, timezone_str)
    match = {"parent_sku":parent_sku}
    if is_duplicate == "true":
        match['_id'] = {"$ne" : ObjectId(product_id)}
    pipeline = [
            {
            "$match": match
            },
            {
            "$lookup": {
                "from": "marketplace",
                "localField": "marketplace_ids",
                "foreignField": "_id",
                "as": "marketplace_ins"
            }
            },
            {
                "$addFields":{
                    "marketplace_ins":{
                        "$cond":{
                            "if":{"$eq":[{"$size":"$marketplace_ins"},0]},
                            "then":[{"name":"Unknown","image_url":""}],
                            'else':"$marketplace_ins"
                        }
                    }
                }
            },
            {
            "$project": {
                "_id": 0,
                "id": {"$toString": "$_id"},
                "product_title": {"$ifNull": ["$product_title", ""]},
                "product_id": {"$ifNull": ["$product_id", ""]},
                "product_id_type": {"$ifNull": ["$product_id_type", ""]},
                "sku": {"$ifNull": ["$sku", ""]},
                "price": {"$ifNull": ["$price", 0]},
                "currency": {"$ifNull": ["$currency", ""]},
                "quantity": {"$ifNull": ["$quantity", 0]},
                "marketplace_ins": {
                "$reduce": {
                    "input": "$marketplace_ins.name",
                    "initialValue": [],
                    "in": {
                    "$cond": {
                        "if": {"$in": ["$$this", "$$value"]},
                        "then": "$$value",
                        "else": {"$concatArrays": ["$$value", ["$$this"]]}
                    }
                    }
                }
                },
                "marketplace_image_url": {
                "$reduce": {
                    "input": "$marketplace_ins.image_url",
                    "initialValue": [],
                    "in": {
                    "$cond": {
                        "if": {"$in": ["$$this", "$$value"]},
                        "then": "$$value",
                        "else": {"$concatArrays": ["$$value", ["$$this"]]}
                    }
                    }
                }
                },
                "brand_name": {"$ifNull": ["$brand_name", ""]},
                "image_url": {"$ifNull": ["$image_url", ""]}
            }
            }
        ]
    variant_list = list(Product.objects.aggregate(*(pipeline)))
    product_ids=[variant['id']for variant in variant_list]
    if not product_ids:
        return variant_list
    sales_data = batch_get_sales_data_optimized(product_ids, start_date, end_date, today_start_date, today_end_date)
    orders_count_raw=OrderItems.objects.aggregate(
            {
                "$match":{
                    "created_date":{"$gte":today_start_date,"$lte":today_end_date},
                    "ProductDetails._id":{"$in":[ObjectId(pid) for pid in product_ids]}
                }
            },
            {
                "$group":{
                    "_id":"$ProductDetails._id",
                    "count":{"$sum":1}
                }
            }
        )
    orders_count={str(item['_id']):item['count'] for item in orders_count_raw}
    for i ,product in enumerate(variant_list):
        pid=product["id"]
        product_id=product['id']
        products_sales=sales_data.get(product_id,{
                "today": {"revenue": 0, "units": 0},
            "period": {"revenue": 0, "units": 0},
            "compare": {"revenue": 0, "units": 0}
            })
        today_revenue=products_sales['today']['revenue']
        period_units=products_sales['period']['units']
        orders_today=orders_count.get(pid,0)
        product.update({
                'salesForToday':round(today_revenue,2),
                'unitsSoldForToday':round(period_units,2),
                "ordersForToday":orders_today
            })
    return clean_json_floats(variant_list)
def backfill_missing_merchant_shipment_cost(batch_size=200):
    print('in backfill`')
    orders = Order.objects(
        merchant_shipment_cost=None,
        order_status__in=["Shipped","Delivered"],
        fulfillment_channel__in=["MFN", "SellerFulfilled"]
    ).order_by("-order_date")[:batch_size]
    for order in orders:
        print(order.id)
    logger.info(f" Found {len(orders)} order(s) to process for backfill.")
    updated_count = 0
    skipped_count = 0
    for order in orders:
        try:
            fc = order.fulfillment_channel
            merchant_id = order.merchant_order_id
            po_id = order.purchase_order_id
            order_date = order.order_date
            email = order.customer_email_id
            if not fc or not order_date:
                logger.warning(f"Skipping order {order.id}: Missing fulfillment_channel or order_date.")
                skipped_count += 1
                continue
            if fc == "MFN":
                if not merchant_id:
                    logger.warning(f"Skipping MFN order {order.id}: Missing merchant_order_id.")
                    skipped_count += 1
                    continue
                logger.info(f"[MFN] Processing order: {merchant_id}")
                order_data = get_full_order_and_shipping_details(merchant_id)
                time.sleep(0.5)
                if order_data and order_data.get("shipments"):
                    latest = order_data["shipments"][-1]
                    cost = float(latest.get("shipmentCost", 0) or 0)
                    if cost > 0:
                        order.update(set__merchant_shipment_cost=cost)
                        updated_count += 1
            elif fc == "SellerFulfilled":
                if not po_id or not email:
                    logger.warning(f"Skipping SellerFulfilled order {order.id}: Missing purchase_order_id or email.")
                    skipped_count += 1
                    continue
                logger.info(f"[Seller] Processing PO: {po_id}")
                shipping_info = get_orders_by_customer_and_date(
                    customer_email=email,
                    order_date_utc_iso=order_date,
                    purchase_order_id=po_id,
                    local_tz='US/Pacific'
                )
                time.sleep(0.5)
                if shipping_info:
                    latest = shipping_info[-1]
                    cost = float(latest.get("shipmentCost", 0) or 0)
                    if cost > 0:
                        order.update(set__merchant_shipment_cost=cost)
                        updated_count += 1
        except Exception as e:
            logger.error(f"Error processing order {order.id}: {e}")
            skipped_count += 1
    logger.info(f"Finished: Updated {updated_count}, Skipped {skipped_count}")
    return {
        "updated": updated_count,
        "skipped": skipped_count,
        "batch_size": batch_size,
        "processed": len(orders)
    }

from django.views.decorators.http import require_http_methods
def clear_cache_by_func_name(func_name, prefix="cache"):
    try:
        if '*' in func_name or '?' in func_name:
            patterns = [func_name]
        elif '.' in func_name:
            patterns = [
                f"{prefix}:{func_name}:*",      
                f"{prefix}:*{func_name}:*",     
            ]
        else:
            patterns = [
                f"{prefix}:*{func_name}:*",     
                f"{prefix}:*.{func_name}:*",    
                f"{prefix}:{func_name}:*",      
            ]
        all_keys_to_delete = set()
        for pattern in patterns:
            try:
                keys = cache.keys(pattern)
                if keys:
                    all_keys_to_delete.update(keys)
                    logger.info(f"Found {len(keys)} keys matching pattern: {pattern}")
            except Exception as e:
                logger.warning(f"Error searching pattern {pattern}: {e}")
                continue
        logger.info(f"Total unique keys found for '{func_name}': {len(all_keys_to_delete)}")
        deleted_count = 0
        failed_count = 0
        for key in all_keys_to_delete:
            try:
                if cache.delete(key):
                    deleted_count += 1
                    logger.debug(f"Deleted cache key: {key}")
                else:
                    failed_count += 1
                    logger.warning(f"Failed to delete cache key: {key}")
            except Exception as e:
                failed_count += 1
                logger.error(f"Error deleting cache key {key}: {e}")
        return deleted_count, failed_count
    except Exception as e:
        logger.error(f"Error in clear_cache_by_func_name: {e}")
        return 0, 0
@require_http_methods(["POST", "GET"])
@csrf_exempt  
def clear_cache_key(request):
    try:
        key = request.GET.get('key') or (request.POST.get('key') if request.method == 'POST' else None)
        clear_all = (request.GET.get('all') or (request.POST.get('all') if request.method == 'POST' else None)) == 'true'
        prefix = request.GET.get('prefix') or (request.POST.get('prefix') if request.method == 'POST' else None) or 'cache'
        if clear_all:
            try:
                cache.clear()
                logger.info("All cache entries cleared")
                return {
                    'status': 'success', 
                    'message': 'All cache entries have been cleared!',
                    'cleared_count': 'all'
                }
            except Exception as e:
                logger.error(f"Error clearing all cache: {e}")
                return {
                    'status': 'error', 
                    'message': f'Error clearing all cache: {str(e)}'
                }
        if key:
            try:
                deleted_count, failed_count = clear_cache_by_func_name(key, prefix)
                if deleted_count > 0:
                    message = f'Cleared {deleted_count} cache entries for "{key}"'
                    if failed_count > 0:
                        message += f' (failed to delete {failed_count} entries)'
                    return {
                        'status': 'success' if failed_count == 0 else 'partial_success',
                        'message': message,
                        'cleared_count': deleted_count,
                        'failed_count': failed_count
                    }
                else:
                    return {
                        'status': 'warning', 
                        'message': f'No cache entries found for "{key}"',
                        'cleared_count': 0
                    }
            except Exception as e:
                logger.error(f"Error clearing cache for key {key}: {e}")
                return {
                    'status': 'error', 
                    'message': f'Error clearing cache: {str(e)}'
                }
        return {
            'status': 'error', 
            'message': 'Provide ?key=function_name (or module.function or pattern) or ?all=true'
        }
    except Exception as e:
        logger.error(f"Unexpected error in clear_cache_key: {e}")
        return {
            'status': 'error', 
            'message': f'Unexpected error: {str(e)}'
        }